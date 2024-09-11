from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill

def image_to_excel(image_path, excel_path):
    # Open the image file
    img = Image.open(image_path)
    img = img.convert('RGB')  # Ensure image is in RGB mode
    
    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Process each pixel and set cell background color
    for y in range(img.height):
        for x in range(img.width):
            r, g, b = img.getpixel((x, y))
            hex_color = f'{r:02X}{g:02X}{b:02X}'
            # Ensure the cell is correctly addressed
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    # Set column widths and row heights to minimal sizes
    for col in range(1, img.width + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2
    for row in range(1, img.height + 1):
        sheet.row_dimensions[row].height = 15

    # Save the workbook
    workbook.save(excel_path)

# Usage
image_path = 'me.jpg'  # Replace with your image path
excel_path = 'output_image.xlsx'       # Replace with your desired output path
image_to_excel(image_path, excel_path)
